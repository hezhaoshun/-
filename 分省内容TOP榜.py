# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""
import pandas as pd
from openpyxl import Workbook
#生成一个播放次数排名的列表
x=pd.read_excel(r"C:\Users\Administrator\Desktop\1 咪视界内容排行走势_日_表1.xlsx")
x1=x[x['分类']=='剔重汇总']
means=x1['播放次数'].groupby(x1['渠道省份']).mean()
mean1=means[(means.index!='剔重汇总')&(means.index!='全国')&(means.index!='未知')]
mean2=list(mean1.sort_values(ascending=False).index[:10])
#案例：将dataframe对象转换成openpyxl方式并输出
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
wb = load_workbook(filename=r"C:\Users\Administrator\Desktop\工作簿1.xlsx")
ws=wb.active
rows = dataframe_to_rows(x1)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
         ws.cell(row=r_idx, column=c_idx, value=value)
wb.save(r"C:\Users\Administrator\Desktop\工作簿1.xlsx")
